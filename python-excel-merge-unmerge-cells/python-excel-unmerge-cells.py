from openpyxl import Workbook, load_workbook

wb = load_workbook(filename = 'excel-merge-cells.xlsx')

ws = wb.active

#ws.unmerge_cells('B2:C2')
#ws.unmerge_cells('B2:C3')

#ws.unmerge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
ws.unmerge_cells(start_row=2, start_column=2, end_row=3, end_column=3)

wb.save('excel-unmerge-cells.xlsx')
