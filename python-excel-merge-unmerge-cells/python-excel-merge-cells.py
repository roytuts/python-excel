from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws.cell(row=2, column=2, value='This is a merge test')
#ws['B2'] = 'This is a merge test'

#ws.merge_cells('B2:C2')
#ws.merge_cells('B2:C3')

#ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
#ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=3)

wb.save('excel-merge-cells.xlsx')
